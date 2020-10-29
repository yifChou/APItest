import aiohttp
import asyncio
import base64
from global_setting import *
import random
import time
from openpyxl import Workbook
import threading
def ramdom_decimal(max,decimal):
    import random
    a = round(random.randint(0,max)+ random.random(),decimal)
    return a
def get_date():
    from datetime import datetime
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    print(date)
    return date
header_test = {
"Host": "120.76.102.19:8034",
"Content-Type": "application/json;charset=UTF-8",
"Content-Length": "1940",
}
def get():
    pass
def get_authorizaiton():
    author = base64.b64encode(str.encode(username + "&" + password))
    au = author.decode()
    #print(au,type(au))
    return au
async def post_basic(u,d):
    auth = aiohttp.BasicAuth(login=username,password=password)
    async with aiohttp.ClientSession(auth=auth) as re:
        response = await re.post(url = u,json = d)
        redata = await response.content.read()
        redata = redata.decode("utf-8")
        await re.close()
        result = redata[redata.find('"ResultDesc":"')+14:redata.find('","Item"')]
        print("result:",result)
        if result == "提交成功":
            a = redata.find('"WayBillNumber":')
            b = redata.find(',"SenderInfoStatus"')
            orderid = redata[a+17:b-1]
            print( redata, "\n",orderid)
            return [1,orderid]
        else:
            #result = redata[redata.find('"Feedback":"') + 12:redata.find('","AgentNumber"')]
            #print(result)
            return [0,result]
#get_authorizaiton()
async def order(orderno,total,length,wide,height,weight):
    order_list = []
    order_fail = []
    wb = Workbook()
    ws = wb.active
    row = ["单号","单号类型","袋牌号","到货时间","长(cm)","宽(cm)","高(cm)","重量(kg)","锁定产品","使用客户重量(是/否)"]
    ws.append(row)
    for i in range(int(total)):
        on = orderno + str(i+1)
        #random.randint(5,30)
        data[0]["OrderNumber"] = on
        data[0]["Weight"] = int(weight)
        data[0]["Length"] = int(length)
        data[0]["Height"] = int(height)
        data[0]["PackageVolume"] = int(wide)
        data[0]["TrackingNumber"] = random.randint(100000000,900000000)
        orderid = await post_basic(u = urls["orderurl"],d= data)
        if orderid[0] !=0:
            order_list.append(orderid[1])
            ws.append([orderid[1], "", "袋子号", "",int(length) , int(wide), int(height), int(weight), "", "是"])
            #time.sleep(int(sleep))
        else:
            order_fail.append(orderid[1])
    #wb.save("C:\\Users\\Administrator\\Desktop\\批量签入签出"+ get_date() + ".xlsx")
    return [order_list,order_fail]
async def order_random(orderno,total,filename):
    def ramdom_decimal(max, decimal):
        import random
        a = round(random.randint(0, max) + random.random(), decimal)
        return a

    def get_date():
        from datetime import datetime
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        print(date)
        return date
    no = ["01","02","03","04","05","06","07","08","09","10",]
    order_list = []
    order_fail = []
    wb = Workbook()
    ws = wb.active
    row = ["单号","单号类型","袋牌号","到货时间","长(cm)","宽(cm)","高(cm)","重量(kg)","锁定产品","使用客户重量(是/否)"]
    ws.append(row)
    try:
        for i in range(int(total)):
            on = orderno + str(i+1)
            #random.randint(5,30)
            data[0]["OrderNumber"] = on
            data[0]["Weight"] = ramdom_decimal(10,3)
            data[0]["Length"] = ramdom_decimal(10,0)
            data[0]["Height"] = ramdom_decimal(10,0)
            data[0]["PackageVolume"] = ramdom_decimal(10,0)
            data[0]["TrackingNumber"] = random.randint(100000000,900000000)
            orderid = await post_basic(u = urls["orderurl"],d= data)
            if orderid[0] != 0:
                order_list.append(orderid[1])
                ws.append([orderid[1], "", "袋子号", "",data[0]["Length"] , data[0]["PackageVolume"], data[0]["Height"], data[0]["Weight"], "", "是"])
                #time.sleep(int(sleep))
            else:
                order_fail.append(orderid[1])
                #time.sleep(int(sleep))
        # wb.save("C:\\Users\\Administrator\\Desktop\\批量签入签出"+ filename + ".xlsx")
        wb.save("doc//批量签入签出" + filename + ".xlsx")
    except Exception as e:
        try:
            # wb.save("C:\\Users\\Administrator\\Desktop\\批量签入签出"+ filename + ".xlsx")
            print("下单失败：",e)
            wb.save("doc//批量签入签出" + filename + ".xlsx")
        except Exception as e:
            #wb.save("C:\\Users\\Administrator\\Desktop\\批量签入签出" + filename + str(random.choice(no)) + ".xlsx")
            print("保存失败：",e)
            wb.save("doc//批量签入签出" + filename +   str(random.choice(no)) + ".xlsx")
    return [order_list,order_fail]
async def order_thread(orderno,i,semaphore):
    on = orderno + str(i + 1)
    print(on)
    async with semaphore:
        # random.randint(5,30)
        data[0]["OrderNumber"] = on
        data[0]["Weight"] = ramdom_decimal(10, 3)
        data[0]["Length"] = ramdom_decimal(10, 0)
        data[0]["Height"] = ramdom_decimal(10, 0)
        data[0]["PackageVolume"] = ramdom_decimal(10, 0)
        data[0]["TrackingNumber"] = random.randint(10000000000, 90000000000)
        orderid = await post_basic(u=urls["orderurl"], d=data)
        print("orderid:",orderid)
        return orderid
async def get_test(url):
    async with aiohttp.ClientSession() as session:
        response = await session.get(url)
        result = await response.read()
        await asyncio.sleep(3)
        #print(result)
        return result
async def aio_get(url):
    result = await get_test(url)
    return result
async def print_page(url):
    header_test["Authorization"] = "Basic {}".format(get_authorizaiton())
    print(header_test)
    data[0]["OrderNumber"] = "201904179000"
    data[0]["Weight"] = ramdom_decimal(10, 3)
    data[0]["Length"] = ramdom_decimal(10, 0)
    data[0]["Height"] = ramdom_decimal(10, 0)
    data[0]["PackageVolume"] = ramdom_decimal(10, 0)
    data[0]["TrackingNumber"] = random.randint(10000000000, 90000000000)
    auth = aiohttp.BasicAuth(login=username,password=password)
    async with aiohttp.ClientSession(auth=auth) as session:# async with用法， ClientSession（）需要使用async with上下文管理器来关闭
        async with session.post(url=urls["orderurl"],json=data) as resp:# post请求
            result =await resp.content.read()
            print(result.decode("utf-8"),resp.status)
            return result.decode("utf-8")
def async_order(ordernum,count,filename):

    start = time.time()
    # tasks = print_page("http://www.baidu.com")
    # tasks = aio_get(url=urls["geturl"])
    #semaphore = asyncio.Semaphore(semaphore)
    semaphore1 = asyncio.Semaphore(3)
    tasks = [asyncio.ensure_future(order_thread(ordernum, i, semaphore=semaphore1)) for i in range(int(count))]
    loop = asyncio.get_event_loop()
    #asyncio.set_event_loop(loop)
    result = loop.run_until_complete(asyncio.gather(*tasks))

    no = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", ]
    order_list = []
    order_fail = []
    wb = Workbook()
    ws = wb.active
    row = ["单号", "单号类型", "袋牌号", "到货时间", "长(cm)", "宽(cm)", "高(cm)", "重量(kg)", "锁定产品", "使用客户重量(是/否)"]
    ws.append(row)
    for orderid in result:
        if orderid[0] != 0:
            order_list.append(orderid[1])
            ws.append([orderid[1], "", "袋子号", "",data[0]["Length"] , data[0]["PackageVolume"], data[0]["Height"], data[0]["Weight"], "", "是"])
            #time.sleep(int(sleep))
        else:
            order_fail.append(orderid[1])
            #time.sleep(int(sleep))
        # wb.save("C:\\Users\\Administrator\\Desktop\\批量签入签出"+ filename + ".xlsx")
    wb.save("doc//批量签入签出" + filename + ".xlsx")
    print("失败请求个数：" + f'{len(order_fail)}' + "/" + f'{len(result)}')
    end = time.time() - start
    print(end)
    return [order_list,order_fail,int(end)]
if __name__ == "__main__":
    filename = "test"
    result = async_order("201904226000",100,filename)
    print(str(result))